
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
from login_zuora import *

#driver = webdriver.Chrome('/Users/davidchen/PycharmProjects/pythonProject/Drivers/chromedriver')
path = "/Users/davidchen/PycharmProjects/pythonProject/Zuora/Datasources/test.xlsx"

#excel data
workbook = openpyxl.load_workbook(path)
sheet = workbook.active

rows = sheet.max_row
cols = sheet.max_column

list = []
for r in range(1,rows+1):
    for c in range(1,cols+1):
        list.append(sheet.cell(row=r,column=c).value)

#navigates to zuora
login()
#driver.get("https://www.zuora.com/apps/CustomerAccount.do?menu=Z-Billing#/page/1/tab/2")

#driver.maximize_window()

#logs into zuora
#username = driver.find_element_by_id("id_username")
#username.clear()
#username.send_keys("david.chen@contentful.com")

#password = driver.find_element_by_name("password")
#password.clear()
#password.send_keys("KLUX3somp-gaur")

#clicks login button
#driver.find_element_by_xpath("/html/body/form[2]/div/a").click()

#BEGIN LOOP
expand = 0
i = 0
len = len(list)
while i <= 9:
    #wait
    time.sleep(3)
    #navigates to zuora
    driver.get("https://www.zuora.com/apps/CustomerAccount.do?menu=Z-Billing#/page/1/tab/2")

    #input Customer Account Number
    driver.find_element_by_xpath("/html/body/div[1]/div/div[1]/div/table/tbody/tr/td[2]/div/div[2]/div/div[2]/div[2]/form/div[1]/input").send_keys(list[i])
    driver.find_element_by_xpath("/html/body/div[1]/div/div[1]/div/table/tbody/tr/td[2]/div/div[2]/div/div[2]/div[2]/form/div[1]/input").send_keys(Keys.ENTER)

    #click Customer Account
    time.sleep(3) #seconds
    driver.find_element_by_xpath("/html/body/div[1]/div/div[1]/div/table/tbody/tr/td[2]/div/div[1]/div/div[3]/div/div/div[2]/div[2]/div/div/table/tbody/tr/td/div[2]/div[1]/div/div/table/tbody/tr[1]/td[1]/a").click()

    #click edit Basic Information
    time.sleep(1)
    driver.find_element_by_xpath("/html/body/div[3]/div/div[1]/div/table/tbody/tr/td[2]/div[4]/div[1]/div/div[1]/div/div[2]/div/div[1]/ul/li/a/span").click()

    #click 'Billing Entity'
    time.sleep(2) #seconds
    driver.find_element_by_xpath("/html/body/div[3]/div/div[1]/div/table/tbody/tr/td[2]/div[4]/div[1]/div/div[1]/div/div/div/div[2]/div/div[2]/div/table/tbody/tr/td[2]/form[1]/table/tbody/tr[17]/td/select").send_keys('inc')
    driver.find_element_by_xpath("/html/body/div[3]/div/div[1]/div/table/tbody/tr/td[2]/div[4]/div[1]/div/div[1]/div/div/div/div[2]/div/div[2]/div/table/tbody/tr/td[2]/form[1]/table/tbody/tr[17]/td/select").send_keys(Keys.ENTER)

    #click 'Save'
    driver.find_element_by_xpath("/html/body/div[3]/div/div[1]/div/table/tbody/tr/td[2]/div[4]/div[1]/div/div[1]/div/div/div/div[2]/div/div[3]/div[1]/span").click()

    #expand Billing and Payment Info
    if expand == 0:
        driver.find_element_by_xpath("/html/body/div[3]/div/div[1]/div/table/tbody/tr/td[2]/div[4]/div[1]/div/div[2]/div/div[2]/div/div[1]/a/h2").click()
        expand += 1

    #click edit Billing and Payment Info
    time.sleep(1)
    driver.find_element_by_xpath("/html/body/div[3]/div/div[1]/div/table/tbody/tr/td[2]/div[4]/div[1]/div/div[2]/div/div[2]/div/div[2]/div[1]/ul/li/a/span").click()

    #click payment gateway
    time.sleep(2) #seconds
    driver.find_element_by_xpath("/html/body/div[3]/div/div[1]/div/table/tbody/tr/td[2]/div[4]/div[1]/div/div[2]/div/div/div/div[2]/div/div[2]/div/form/table/tbody/tr[6]/td/select").send_keys("Stripe Gateway Inc 3DSv2")
    driver.find_element_by_xpath("/html/body/div[3]/div/div[1]/div/table/tbody/tr/td[2]/div[4]/div[1]/div/div[2]/div/div/div/div[2]/div/div[2]/div/form/table/tbody/tr[6]/td/select").send_keys(Keys.ENTER)

    #click Invoice Template
    driver.find_element_by_xpath("/html/body/div[3]/div/div[1]/div/table/tbody/tr/td[2]/div[4]/div[1]/div/div[2]/div/div/div/div[2]/div/div[2]/div/form/table/tbody/tr[12]/td/select").send_keys("INC SS Template (January 2017)")
    driver.find_element_by_xpath("/html/body/div[3]/div/div[1]/div/table/tbody/tr/td[2]/div[4]/div[1]/div/div[2]/div/div/div/div[2]/div/div[2]/div/form/table/tbody/tr[12]/td/select").send_keys(Keys.ENTER)

    #click Save
    driver.find_element_by_xpath("/html/body/div[3]/div/div[1]/div/table/tbody/tr/td[2]/div[4]/div[1]/div/div[2]/div/div/div/div[2]/div/div[3]/div[1]/span").click()

    #navigate back to customer account page
    print('Account', list[i], 'has been updated')
    i += 1
    driver.get("https://www.zuora.com/apps/CustomerAccount.do?menu=Z-Billing#/page/1/tab/2")

