from selenium import webdriver
import time
import openpyxl

driver = webdriver.Chrome('/Users/davidchen/PycharmProjects/pythonProject/Drivers/chromedriver')
path = "/Users/davidchen/PycharmProjects/pythonProject/Zuora/Datasources/ps.xlsx"

#excel data
workbook = openpyxl.load_workbook(path)
sheet = workbook.active

rows = sheet.max_row
cols = sheet.max_column

#log into Zuora
driver.get("https://www.zuora.com/apps/CustomerAccount.do?menu=Z-Billing#/page/1/tab/2")
driver.maximize_window()
time.sleep(1)

username = driver.find_element_by_id("id_username")
username.clear()
username.send_keys("david.chen@contentful.com")

password = driver.find_element_by_name("password")
password.clear()
password.send_keys("KLUX3somp-gaur")

#clicks login button
driver.find_element_by_xpath("/html/body/form[2]/div/a").click()
time.sleep(5)



# BEGIN LOOP

for r in range(1,rows+1):

    #loop list creation for row data
    list = []
    for c in range(1,cols+1):
        list.append(sheet.cell(row=r,column=c).value)
    print(list)

    #go to account URL
    driver.get(list[4])
    time.sleep(2)

    #verify correct subscription
    if list[3] == driver.find_element_by_xpath("/html/body/div[3]/div/div[1]/div/table/tbody/tr/td[2]/div[4]/div[1]/div/div[4]/div[2]/div/div[4]/div[1]/div[2]/table/tbody/tr[2]/td[2]/div/a[2]").text:
        driver.find_element_by_xpath("/html/body/div[3]/div/div[1]/div/table/tbody/tr/td[2]/div[4]/div[1]/div/div[4]/div[2]/div/div[4]/div[1]/div[2]/table/tbody/tr[2]/td[2]/div/a[2]").click()
    else:
        print("Subscription does not match")
        quit()

    #click 'amendment' on subscription page
    time.sleep(1)
    if driver.find_element_by_xpath("/html/body/div[1]/div/div[1]/div/table/tbody/tr/td[2]/div[2]/ul/li[3]/a/span").text == 'amendment':
        driver.find_element_by_xpath("/html/body/div[1]/div/div[1]/div/table/tbody/tr/td[2]/div[2]/ul/li[3]/a/span").click()
    elif driver.find_element_by_xpath("/html/body/div[1]/div/div[1]/div/table/tbody/tr/td[2]/div[2]/ul/li[2]/a/span").text == 'amendment':
        driver.find_element_by_xpath("/html/body/div[1]/div/div[1]/div/table/tbody/tr/td[2]/div[2]/ul/li[2]/a/span").click()
    else:
        print('Cannot find amendment button')
        quit()

    #fill out amendment information
        ##amendment name
    driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/div/div/div/div/div/dl/dd/form/div[1]/div/table/tbody/tr[6]/td[3]/input").send_keys(list[0])

        ##change description
    driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/div/div/div/div/div/dl/dd/form/div[1]/div/table/tbody/tr[10]/td[3]/textarea").send_keys("Effective ", list[1])

        ##amendment type
    driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/div/div/div/div/div/dl/dd/form/div[1]/div/table/tbody/tr[13]/td[3]/select").send_keys("New Product")
    time.sleep(1)

        ##save details
    driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/div/div/div/div/div/dl/dd/form/div[2]/div[1]/span").click()

    #add new product & rate plan
    driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/div[3]/div/div/div/dl/dt/a").click()
    time.sleep(1)

        ##input Professional Services
    driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/div[3]/div/div/div/dl/dd/div/div/div[1]/div[2]/div[1]/input[1]").send_keys("Professional Services")
    driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/div[3]/div/div/div/dl/dt/a").click()
    time.sleep(2)

    #select PS type
    if list[0] == "Premium":
        #select Premium
        driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/div[3]/div/div/div/dl/dd/div/div/div[1]/div[3]/table/tbody/tr[1]/td[1]/a").click()
    elif list[0] == "Quickstart":
        #select Quickstart
        driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/div[3]/div/div/div/dl/dd/div/div/div[1]/div[3]/table/tbody/tr[2]/td[1]/a").click()
    elif list[0] == "TAM":
        #select TAM
        driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/div[3]/div/div/div/dl/dd/div/div/div[1]/div[3]/table/tbody/tr[4]/td[1]/a").click()
    elif list[0] == "Intro":
        #select Contentful Accelerator - Intro to Contentful
        driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/div[3]/div/div/div/dl/dd/div/div/div[1]/div[3]/table/tbody/tr[3]/td[1]/a").click()
    elif list[0] == "Content Modeling":
        #selecct Contentful Accelerator - Content Modeling
        driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/div[3]/div/div/div/dl/dd/div/div/div[1]/div[3]/table/tbody/tr[3]/td[1]/a").click()
    else:
        print('ERROR: PS product could not be found')

    #expand PS product
    time.sleep(2)
    driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/div[3]/div/div/div/dl/dd/div/div/div[2]/div/table/tbody/tr[3]/td[1]/a").click()

    #update PS price
    time.sleep(2)
    driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/div[3]/div/div/div/dl/dd/div/div/div[2]/div/table/tbody/tr[3]/td/form/div/table/tbody/tr[9]/td[2]/input[2]").clear()
    driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/div[3]/div/div/div/dl/dd/div/div/div[2]/div/table/tbody/tr[3]/td/form/div/table/tbody/tr[9]/td[2]/input[2]").send_keys(list[2])

    #save changes
    driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/div[3]/div/div/div/dl/dd/div/div/div[2]/div/table/tbody/tr[3]/td/form/div/table/tbody/tr[13]/td/div/div[1]/span").click()
    time.sleep(1)

    #update contract effective
    driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/div[2]/div/div/div/div/dl/dt/a[4]/p/span").click()
    time.sleep(1)
    driver.find_element_by_xpath("/html/body/div[8]/div[2]/div/div/dl/dd/table/tbody/tr[4]/td[2]/input").send_keys(list[1])
    driver.find_element_by_xpath("/html/body/div[8]/div[2]/div/div/dl/dd/table/tbody/tr[6]/td/div/div[1]/span").click()
    time.sleep(2)

    #update service activation
    driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/div[2]/div/div/div/div/dl/dt/a[3]/p/span").click()
    time.sleep(1)
    driver.find_element_by_xpath("/html/body/div[8]/div[2]/div/div/dl/dd/table/tbody/tr[5]/td[2]/input").send_keys(list[1])
    driver.find_element_by_xpath("/html/body/div[8]/div[2]/div/div/dl/dd/table/tbody/tr[7]/td/div/div[1]/span").click()
    time.sleep(2)

    print("Row", r, "has successfully been provisioned with the list above")



